"""
Módulo de generación de Anexos C y D en formato Excel para el trámite SGDA
(Sistema de Generación Distribuida para Autoabastecimiento) ante la EEQ.

Anexo C (Registro Histórico de Consumos):
  Tabla con el consumo mensual real de los últimos 12 períodos.
  Se usa cuando hay 12 o más facturas válidas.
  Compatible con clientes residenciales e industriales (en ambos casos se muestra
  el consumo total mensual sin desglose de franjas horarias).

Anexo D (Estimación de Consumo por Aparatos Eléctricos):
  Tabla de 21 aparatos domésticos con potencias, factores y horas de uso,
  escalados proporcionalmente al consumo anual real del cliente.
  Se usa cuando hay menos de 12 facturas válidas (nueva instalación o datos
  insuficientes para el Anexo C).
  Para clientes industriales el consumo de escala es la suma de todas las
  franjas horarias (consumo_total_kwh), igual que para residencial.

Los templates base de los aparatos se leen del archivo AO_Macro.xlsm.
"""

import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR   = Path(__file__).parent
TEMPLATE   = BASE_DIR / "AO_Macro.xlsm"
CONSUMO_BASE_ANEXO_D = 14747.28  # kWh/año del template AO_Macro.xlsm
MAX_HORAS  = 8760                 # Máximo de horas en un año

MESES_ES = {
    1: "Ene", 2: "Feb",  3: "Mar", 4: "Abr",
    5: "May", 6: "Jun",  7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}

# Horas de uso anuales base para cada aparato (columna J, filas 11-31 del template)
_HORAS_BASE: dict[int, int] = {
    11: 144,   # Licuadora
    12: 360,   # Cafetera
    13: 180,   # Microonda
    14: 720,   # Cocina Eléctrica
    15: 36,    # Tostadora
    16: 720,   # Lavadora
    17: 720,   # Secadora
    18: 360,   # Plancha
    19: 2880,  # TV LCD
    20: 2880,  # Equipo de Cable
    21: 2880,  # Consola Videojuegos
    22: 1440,  # Equipo de Sonido
    23: 1800,  # Computadora Portátil
    24: 8640,  # Modem
    25: 8640,  # Enrutador
    26: 8760,  # Refrigeradora
    27: 365,   # Ducha eléctrica
    28: 365,   # Impresora
    29: 2190,  # Ventilador
    30: 730,   # Aspiradora
    31: 3600,  # Bombillas LED
}


def _cargar_template_aparatos() -> list[dict]:
    """
    Lee los datos base de aparatos del template AO_Macro.xlsm.

    Lee las filas 11 a 31 de la hoja 'Anexo D' con las columnas:
      B = nombre, C = cantidad, D = potencia nominal (kW),
      F = factor de frecuencia de uso, H = factor de simultaneidad.

    Retorna:
        Lista de dicts con los datos de cada aparato, o lista vacía si el
        template no existe o no se puede leer.
    """
    if not TEMPLATE.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(TEMPLATE), keep_vba=True, data_only=True)
        ws = wb["Anexo D"]
        aparatos = []
        for row_idx in range(11, 32):
            nombre   = ws.cell(row=row_idx, column=2).value
            cantidad = ws.cell(row=row_idx, column=3).value
            pot_kw   = ws.cell(row=row_idx, column=4).value
            freq     = ws.cell(row=row_idx, column=6).value
            simul    = ws.cell(row=row_idx, column=8).value
            horas_base = _HORAS_BASE.get(row_idx, 0)
            if nombre and cantidad is not None and pot_kw is not None:
                aparatos.append({
                    "item":              row_idx - 10,
                    "nombre":            str(nombre),
                    "cantidad":          float(cantidad) if cantidad else 0,
                    "potencia_kw":       float(pot_kw) if pot_kw else 0,
                    "factor_frecuencia": float(freq) if freq is not None else 0.8,
                    "factor_simul":      float(simul) if simul is not None else 1.0,
                    "horas_base":        horas_base,
                })
        return aparatos
    except Exception:
        return []


# Datos de aparatos cacheados al importar el módulo para no releer el archivo
_APARATOS_BASE: list[dict] = _cargar_template_aparatos()


def _mes_anio_de(fecha_str: str) -> tuple[int, int]:
    """
    Extrae (mes, año) de una fecha en formato 'DD-MM-YYYY'.

    Parámetros:
        fecha_str: Cadena de fecha.

    Retorna:
        Tupla (mes, año).
    """
    d = datetime.strptime(fecha_str, "%d-%m-%Y")
    return (d.month, d.year)


# ---------------------------------------------------------------------------
# Persistencia JSON
# ---------------------------------------------------------------------------

def generar_json(resultado: dict, ruta: Path) -> Path:
    """
    Guarda el resultado del análisis en un archivo JSON.

    Crea los directorios padre si no existen.

    Parámetros:
        resultado: Diccionario con el resultado del análisis.
        ruta: Ruta destino del archivo .json.

    Retorna:
        Ruta del archivo guardado.
    """
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as fp:
        json.dump(resultado, fp, ensure_ascii=False, indent=2, default=str)
    return ruta


# ---------------------------------------------------------------------------
# Helpers de estilo Excel
# ---------------------------------------------------------------------------

def _header_cell(cell, valor, color_hex: str, bold: bool = True, align: str = "center"):
    """
    Aplica estilo de celda de encabezado: fondo de color, texto blanco y negrita.

    Parámetros:
        cell: Celda de openpyxl a estilizar.
        valor: Valor a escribir en la celda.
        color_hex: Color de fondo en formato hexadecimal (sin '#').
        bold: Si el texto debe ser negrita.
        align: Alineación horizontal del texto.
    """
    cell.value     = valor
    cell.font      = Font(bold=bold, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=color_hex)
    cell.alignment = Alignment(horizontal=align, vertical="center")


# ---------------------------------------------------------------------------
# Anexo C — Registro histórico de consumos
# ---------------------------------------------------------------------------

def generar_excel_anexo_c(resultado: dict, ruta: Path) -> Path:
    """
    Crea un archivo Excel (.xlsx) con la tabla del Anexo C.

    La tabla muestra el consumo mensual en kWh para cada período facturado.
    Incluye una fila de total al final.

    Compatible con clientes residenciales e industriales: en ambos casos
    se muestra el consumo total mensual (consumo_total_kwh por período).

    Acepta facturas con valores de consumo_total_kwh ya editados por el usuario
    (si el frontend envió valores modificados).

    Parámetros:
        resultado: Diccionario de análisis con clave 'facturas' (lista de facturas
                   válidas ordenadas por fecha).
        ruta: Ruta destino del archivo Excel.

    Retorna:
        Ruta del archivo Excel generado.
    """
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anexo C"

    facturas = resultado.get("facturas", [])
    tiene_demanda = any(f.get("valor_demanda") for f in facturas)

    # Determinar columnas según tipo de cliente
    if tiene_demanda:
        cols = ["#", "Fecha (Mes/Año)", "Consumo (kWh)", "Valor Planilla (USD)", "Valor Demanda (USD)"]
        merge_rng = "A1:E1"
    else:
        cols = ["#", "Fecha (Mes/Año)", "Consumo (kWh)", "Valor Planilla (USD)"]
        merge_rng = "A1:D1"

    # Título principal
    ws.merge_cells(merge_rng)
    _header_cell(ws["A1"], "REGISTRO HISTÓRICO DE CONSUMOS (kWh)", "1B3A6B", bold=True)
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")

    # Encabezados de columna
    for col_letter, header in zip("ABCDE", cols):
        _header_cell(ws[f"{col_letter}2"], header, "1E4080")

    # Filas de datos
    for i, f in enumerate(facturas):
        row = i + 3
        mes, anio = _mes_anio_de(f["fecha_hasta"])
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = f"{MESES_ES[mes]} {anio}"
        ws[f"C{row}"] = round(f["consumo_total_kwh"], 2)
        ws[f"D{row}"] = round(f.get("valor_total_planilla") or 0, 2)
        if tiene_demanda:
            ws[f"E{row}"] = round(f.get("valor_demanda") or 0, 2)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].number_format = "#,##0.00"
        ws[f"D{row}"].number_format = "#,##0.00"
        if tiene_demanda:
            ws[f"E{row}"].number_format = "#,##0.00"

    # Fila de total
    total_row = len(facturas) + 3
    ws[f"B{total_row}"] = "Total"
    ws[f"B{total_row}"].font      = Font(bold=True)
    ws[f"B{total_row}"].alignment = Alignment(horizontal="right")
    total_kwh = round(sum(f["consumo_total_kwh"] for f in facturas), 2)
    ws[f"C{total_row}"] = total_kwh
    ws[f"C{total_row}"].font          = Font(bold=True)
    ws[f"C{total_row}"].number_format = "#,##0.00"
    total_planilla = round(sum(f.get("valor_total_planilla") or 0 for f in facturas), 2)
    ws[f"D{total_row}"] = total_planilla
    ws[f"D{total_row}"].font          = Font(bold=True)
    ws[f"D{total_row}"].number_format = "#,##0.00"
    if tiene_demanda:
        total_demanda = round(sum(f.get("valor_demanda") or 0 for f in facturas), 2)
        ws[f"E{total_row}"] = total_demanda
        ws[f"E{total_row}"].font          = Font(bold=True)
        ws[f"E{total_row}"].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    if tiene_demanda:
        ws.column_dimensions["E"].width = 20

    wb.save(str(ruta))
    return ruta


# ---------------------------------------------------------------------------
# Anexo D — Estimación de consumo por aparatos eléctricos
# ---------------------------------------------------------------------------

def _derivar_fila(f: dict) -> dict:
    """
    Recalcula las columnas derivadas de una fila del Anexo D a partir de
    los valores base (cantidad, potencia, factores y horas).

    Fórmulas:
      pot_total     = cantidad × potencia_nominal
      carga_inst    = pot_total × factor_frecuencia
      demanda_max   = pot_total × factor_simultaneidad
      energia_kwh   = carga_inst × factor_simultaneidad × horas_anuales

    Parámetros:
        f: Diccionario con campos base de la fila.

    Retorna:
        Diccionario con todos los campos base más los derivados.
    """
    cant  = f.get("cantidad", 0) or 0
    pot   = f.get("potencia_kw", 0) or 0
    freq  = f.get("factor_frecuencia", 0.8) or 0.8
    simul = f.get("factor_simul", 1.0) or 1.0
    horas = f.get("horas_anio", 0) or 0
    pot_total = round(cant * pot, 3)
    carga     = round(pot_total * freq, 3)
    demanda   = round(pot_total * simul, 3)
    energia   = round(carga * simul * horas, 2)
    return {**f, "pot_total_kw": pot_total, "carga_instalada_kw": carga,
            "demanda_max_kw": demanda, "energia_kwh": energia}


def _calcular_filas_anexo_d(resultado: dict) -> tuple[list[dict], list[str]]:
    """
    Escala las horas de uso de cada aparato proporcionalmente al consumo anual
    real del cliente y calcula todas las columnas derivadas.

    El factor de escala es: consumo_real / consumo_base_template.
    Las horas de cada aparato se multiplican por este factor, con un tope de
    MAX_HORAS (8760 h/año).

    Para clientes industriales, 'consumo_real' es la suma de todas las franjas
    horarias (consumo_anual_kwh o consumo_anual_estimado_kwh).

    Parámetros:
        resultado: Diccionario del análisis con 'consumo_anual_kwh' o
                   'consumo_anual_estimado_kwh'.

    Retorna:
        Tupla (lista de filas con valores calculados, lista de advertencias).
    """
    consumo = (
        resultado.get("consumo_anual_estimado_kwh")
        or resultado.get("consumo_anual_kwh")
        or CONSUMO_BASE_ANEXO_D
    )

    # Consumo base calculado con la fórmula completa (incluyendo freq y simul)
    total_base = sum(
        ap["cantidad"] * ap["potencia_kw"] * ap["factor_frecuencia"] * ap["factor_simul"] * ap["horas_base"]
        for ap in _APARATOS_BASE
    ) or CONSUMO_BASE_ANEXO_D

    factor = consumo / total_base

    advertencias: list[str] = []
    filas: list[dict] = []

    for ap in _APARATOS_BASE:
        horas_scaled = min(round(ap["horas_base"] * factor), MAX_HORAS)

        if horas_scaled >= MAX_HORAS and ap["horas_base"] < MAX_HORAS:
            advertencias.append(
                f"{ap['nombre']}: horas limitadas a 8760 h/año "
                f"(escala {factor:.2f}× superó el máximo)"
            )

        fila_base = {
            "item":              ap["item"],
            "nombre":            ap["nombre"],
            "cantidad":          ap["cantidad"],
            "potencia_kw":       ap["potencia_kw"],
            "factor_frecuencia": ap["factor_frecuencia"],
            "factor_simul":      ap["factor_simul"],
            "horas_anio":        horas_scaled,
        }
        filas.append(_derivar_fila(fila_base))

    return filas, advertencias


def tabla_anexo_d(resultado: dict) -> list[dict]:
    """
    Retorna la lista de filas del Anexo D para incluir en la respuesta de la API.

    Parámetros:
        resultado: Diccionario del análisis.

    Retorna:
        Lista de dicts, uno por aparato, con todos los campos calculados.
    """
    filas, _ = _calcular_filas_anexo_d(resultado)
    return filas


def generar_excel_anexo_d(resultado: dict, ruta: Path, filas_override=None) -> tuple[Path, list[str]]:
    """
    Crea un archivo Excel (.xlsx) con la tabla del Anexo D.

    Si se pasan filas_override, se usan esas filas (valores editados por el usuario
    desde el frontend) en lugar de calcular desde el consumo del resultado.

    La tabla incluye 11 columnas: #, Electrodoméstico, Cant., Pot. nom. (kW),
    Pot. total (kW), F. Frecuencia, Carga inst. (kW), F. Simultaneidad,
    Dem. máx. (kW), Horas/año, Consumo (kWh/año).

    Parámetros:
        resultado: Diccionario del análisis (para calcular si no hay override).
        ruta: Ruta destino del archivo Excel.
        filas_override: Lista de filas con valores a usar directamente (opcional).

    Retorna:
        Tupla (ruta del archivo generado, lista de advertencias).
    """
    ruta = ruta.with_suffix(".xlsx")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    if filas_override:
        filas = [_derivar_fila(f) for f in filas_override]
        advertencias = []
    else:
        filas, advertencias = _calcular_filas_anexo_d(resultado)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anexo D"

    COLOR_HDR  = "1B3A6B"
    COLOR_COL  = "1E4080"
    COLOR_INFO = "D9E1F2"  # Azul claro para las celdas de valor del encabezado

    # ── Encabezado de información del cliente (filas 1-7) ──────────────────
    tipo_label = (
        "Industrial"
        if resultado.get("tipo_cliente", "residencial") == "industrial"
        else "Residencial"
    )
    info_filas = [
        ("Nombre solicitante:",    resultado.get("razon_social", "")),
        ("CI/RUC solicitante:",    resultado.get("cuenta_contrato", "")),
        ("Nombre Rep. Técnico:",   ""),
        ("CI/RUC Rep. Técnico:",   ""),
        ("Localización:",          ""),
        ("Tipo de usuario:",       tipo_label),
        ("Fecha de elaboración:",  datetime.today().strftime("%d/%m/%Y")),
    ]
    for fila_idx, (etiqueta, valor) in enumerate(info_filas, 1):
        celda_etiq = ws.cell(row=fila_idx, column=1)
        celda_etiq.value     = etiqueta
        celda_etiq.font      = Font(bold=True)
        celda_etiq.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(
            start_row=fila_idx, start_column=2,
            end_row=fila_idx,   end_column=11,
        )
        celda_val = ws.cell(row=fila_idx, column=2)
        celda_val.value     = valor
        celda_val.fill      = PatternFill("solid", fgColor=COLOR_INFO)
        celda_val.alignment = Alignment(horizontal="left", vertical="center")
        celda_val.font      = Font(bold=bool(valor))

    # ── Fila 8 en blanco (separador) ───────────────────────────────────────

    # ── Título principal (fila 9, ocupa las 11 columnas) ──────────────────
    ws.merge_cells("A9:K9")
    _header_cell(ws["A9"], "ESTIMACIÓN DE CONSUMO POR APARATOS ELÉCTRICOS", COLOR_HDR, bold=True)
    ws["A9"].font = Font(bold=True, size=12, color="FFFFFF")

    # ── Encabezados de columna en 2 filas (filas 10-11) ───────────────────
    # Fila 10: encabezados de grupo (algunos con rowspan 2)
    ws.merge_cells("A10:A11")   # ITEM
    _header_cell(ws["A10"], "ITEM", COLOR_COL)
    ws.cell(row=10, column=1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws.merge_cells("B10:D10")   # APARATOS ELÉCTRICOS Y DE ALUMBRADO (grupo)
    _header_cell(ws["B10"], "APARATOS ELÉCTRICOS Y DE ALUMBRADO", COLOR_COL)

    for col, texto in (
        (5,  "Potencia Total\nInstalada (kW)"),
        (6,  "Factor de\nFrecuencia de Uso"),
        (7,  "Carga Instalada\ndel Consumidor (kW)"),
        (8,  "Factor de\nSimultaneidad"),
        (9,  "Demanda Máxima\nUnitaria (kW)"),
        (10, "Horas de Uso\nAnual (h/año)"),
        (11, "Consumo\nAnual (kWh)"),
    ):
        ws.merge_cells(
            start_row=10, start_column=col,
            end_row=11,   end_column=col,
        )
        celda = ws.cell(row=10, column=col)
        _header_cell(celda, texto, COLOR_COL)
        celda.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Fila 11: sub-encabezados del grupo APARATOS
    for col, texto in ((2, "DESCRIPCIÓN"), (3, "CANTIDAD"), (4, "Potencia\nnominal (kW)")):
        celda = ws.cell(row=11, column=col)
        _header_cell(celda, texto, COLOR_COL)
        celda.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    ws.row_dimensions[10].height = 30
    ws.row_dimensions[11].height = 28

    # ── Filas de datos (desde fila 12) ────────────────────────────────────
    fmt_3 = "#,##0.000"
    fmt_2 = "#,##0.00"
    fmt_0 = "#,##0"
    DATA_INICIO = 12

    for i, f in enumerate(filas):
        row = DATA_INICIO + i
        ws.cell(row=row, column=1).value  = f["item"]
        ws.cell(row=row, column=2).value  = f["nombre"]
        ws.cell(row=row, column=3).value  = f["cantidad"]
        ws.cell(row=row, column=4).value  = f["potencia_kw"]
        ws.cell(row=row, column=5).value  = f.get("pot_total_kw", 0)
        ws.cell(row=row, column=6).value  = f.get("factor_frecuencia", 0)
        ws.cell(row=row, column=7).value  = f.get("carga_instalada_kw", 0)
        ws.cell(row=row, column=8).value  = f.get("factor_simul", 0)
        ws.cell(row=row, column=9).value  = f.get("demanda_max_kw", 0)
        ws.cell(row=row, column=10).value = f["horas_anio"]
        ws.cell(row=row, column=11).value = f["energia_kwh"]

        ws.cell(row=row, column=1).alignment  = Alignment(horizontal="center")
        ws.cell(row=row, column=3).alignment  = Alignment(horizontal="center")
        for c in (4, 5, 7, 9): ws.cell(row=row, column=c).number_format = fmt_3
        for c in (6, 8):        ws.cell(row=row, column=c).number_format = fmt_2
        ws.cell(row=row, column=10).number_format = fmt_0
        ws.cell(row=row, column=11).number_format = fmt_2

    # ── Fila de total ──────────────────────────────────────────────────────
    total_row = DATA_INICIO + len(filas)
    ws.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row,   end_column=10,
    )
    celda_total_lbl = ws.cell(row=total_row, column=1)
    celda_total_lbl.value     = "CONSUMO ANUAL TOTAL (kWh/año)"
    celda_total_lbl.font      = Font(bold=True)
    celda_total_lbl.alignment = Alignment(horizontal="right", vertical="center")

    total_energia = round(sum(f["energia_kwh"] for f in filas), 2)
    celda_total_val = ws.cell(row=total_row, column=11)
    celda_total_val.value          = total_energia
    celda_total_val.font           = Font(bold=True)
    celda_total_val.number_format  = fmt_2

    # ── Anchos de columna ──────────────────────────────────────────────────
    # Columna A más ancha para acomodar las etiquetas del encabezado (filas 1-7)
    for col_letter, width in zip("ABCDEFGHIJK", [24, 28, 7, 14, 14, 13, 14, 16, 14, 11, 18]):
        ws.column_dimensions[col_letter].width = width

    # Ajustar alto de filas de encabezado de cliente
    for r in range(1, 8):
        ws.row_dimensions[r].height = 18

    wb.save(str(ruta))
    return ruta, advertencias


# ---------------------------------------------------------------------------
# Detalle de planilla (franjas horarias, reactiva, demandas) — tabla aparte
# ---------------------------------------------------------------------------

def generar_excel_detalle_planilla(filas: list[dict], ruta: Path) -> Path:
    """
    Crea un archivo Excel (.xlsx) con el detalle de planilla de una factura
    con tarifa horaria diferenciada (franjas de energía, reactiva, demandas).

    Solo incluye Descripción y Consumo Total, según lo solicitado — el resto
    de columnas (fechas, lecturas, unidad, monto) se muestran únicamente en
    la tabla de la interfaz.

    Parámetros:
        filas: Lista de dicts con al menos 'descripcion' y 'consumo_total'.
        ruta: Ruta destino del archivo Excel.

    Retorna:
        Ruta del archivo Excel generado.
    """
    ruta = ruta.with_suffix(".xlsx")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Planilla"

    ws.merge_cells("A1:B1")
    _header_cell(ws["A1"], "DETALLE DE PLANILLA", "1B3A6B", bold=True)
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")

    _header_cell(ws["A2"], "Descripción", "1E4080")
    _header_cell(ws["B2"], "Consumo Total", "1E4080")

    for i, f in enumerate(filas):
        row = i + 3
        ws[f"A{row}"] = f.get("descripcion", "")
        ws[f"B{row}"] = round(f.get("consumo_total") or 0, 2)
        ws[f"B{row}"].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 16

    wb.save(str(ruta))
    return ruta
